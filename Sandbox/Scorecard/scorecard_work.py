#Import dependencies:
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook


def read_ppd(ppd_file_location):
    ppd = pd.read_csv(ppd_file_location)

    #Nones
    ppd['ADDRESS_UNDELIVERABLE_FLAG_2']=ppd['ADDRESS_UNDELIVERABLE_FLAG'].fillna('None')
    ppd['PRESUMED_DEAD_FLAG_2']=ppd['PRESUMED_DEAD_FLAG'].fillna('None')

    #Everybody gotta be alive
    ppd=ppd[ppd['PRESUMED_DEAD_FLAG_2']=='None']

    #Define subset of DOs
    sub_ppd=ppd[ppd['TOP_CD']=='020']

    #Address Type Table
    # address_type = all_deliverable.groupby('ADDRESS_TYPE').count()[0:1]
    return(ppd, sub_ppd)

#Total
def get_total_table(ppd_df):
    all_deliverable = ppd_df[ppd_df['ADDRESS_UNDELIVERABLE_FLAG_2']=='None']
    counts = ppd_df.count()
    fax_count = counts['FAXNUMBER']
    polo_count = counts['POLO_MAILING_LINE_2']
    telephone_count = counts['TELEPHONE_NUMBER']
    no_delivery_count = counts['ADDRESS_UNDELIVERABLE_FLAG']
    deliverable_counts = all_deliverable.count()
    mailing_address_count=deliverable_counts['MAILING_LINE_2']
    top_count = len(ppd_df[ppd_df['TOP_CD']!=100])
    pe_cd_count = len(ppd_df[ppd_df['PE_CD']!=110])
    prim_spec_count = len(ppd_df[ppd_df['PRIM_SPEC_CD']!='US'])
    total_records=len(ppd_df)

    value_dict={
        'totalrecords':total_records,
        'mailingaddress':mailing_address_count,
        'polo':polo_count,
        'telephone':telephone_count,
        'faxnumber':fax_count,
        'top':top_count,
        'pe':pe_cd_count,
        'primspec':prim_spec_count
    }

    transposed = pd.DataFrame(value_dict, ['Count']).transpose()

    return(transposed)

def get_wslive_results(wslive_file_location):

    wslive_table = pd.read_excel('y-WSLive-Results-20200106.xlsm', sheet_name = 'Summary Percentage', header=3).dropna()
    
    values = list(wslive_table[wslive_table['POLO Address Status']=='Confirmed'].iloc[:,-1])
    values.remove(1)

    col_list = ['polo_correct',
    'telephone_correct',
    'telephone_2_correct',
    'fax_correct',
    'employment_correct']

    wslive_df = pd.DataFrame({'1':col_list,'2':values})

    return(wslive_df)


def update_history_notes(history_worksheet):
    history_list =[]
    for row in history_worksheet.iter_rows(min_row=1, min_col = 3, max_col=5, max_row=20, values_only=True):
        history_list.append(row)
    row_num=1
    for row in history_list:
        history_worksheet['B'+str(row_num)]=row[0]
        history_worksheet['C'+str(row_num)]=row[1]
        history_worksheet['D'+str(row_num)]=row[2]
        row_num+=1
    return(history_worksheet)

def update_notes(notes_worksheet):
    

def read_old_scorecard(old_scorecard_location):
    wb = load_workbook(old_scorecard_location)
    notes_worksheet = wb['Notes']
    history_workhseet = wb['Notes_2']

    return(notes_worksheet, history_worksheet)